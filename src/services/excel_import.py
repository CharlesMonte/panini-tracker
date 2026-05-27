from __future__ import annotations

import math
import re
from dataclasses import dataclass, field
from pathlib import Path
from tempfile import NamedTemporaryFile
from typing import BinaryIO

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from sqlalchemy import select
from sqlalchemy.orm import Session

from src.models import ActionLog, Holding, ImportRun, Person, Sticker
from src.repositories import get_or_create_holding, upsert_person, upsert_sticker
from src.utils.normalization import make_display_code, normalize_person_name, normalize_sticker_code, split_sticker_code


IGNORED_HEADER_PATTERNS = [
    "cherchent",
    "double",
    "echange",
    "échange",
    "detail",
    "détail",
    "possible",
    "sticker total",
    "proportion",
]


@dataclass
class ParsedStickerRow:
    album_order: int
    raw_category: str
    sticker_code: str
    display_code: str
    category_code: str | None
    sticker_number: int | None
    quantities: dict[str, int]


@dataclass
class ExcelPreview:
    sheet_name: str
    header_row: int
    category_column: int
    number_column: int
    person_columns: dict[int, str]
    rows: list[ParsedStickerRow] = field(default_factory=list)
    ignored_rows: list[dict] = field(default_factory=list)

    @property
    def sticker_count(self) -> int:
        return len(self.rows)

    @property
    def people_count(self) -> int:
        return len(self.person_columns)

    @property
    def people_names(self) -> list[str]:
        return list(self.person_columns.values())


def _clean_header(value: object) -> str:
    return " ".join(str(value or "").strip().split())


def _is_ignored_header(value: str) -> bool:
    normalized = value.lower()
    return any(pattern in normalized for pattern in IGNORED_HEADER_PATTERNS)


def _is_formula_cell(cell: Cell) -> bool:
    return isinstance(cell.value, str) and cell.value.startswith("=")


def _as_quantity(value: object) -> int:
    if value is None or value == "":
        return 0
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return 0
        value = value.replace(",", ".")
    try:
        numeric = float(value)
    except (TypeError, ValueError):
        return 0
    if math.isnan(numeric) or numeric < 0:
        return 0
    return int(numeric)


def _number_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    return re.sub(r"\.0$", "", text)


def _find_header_row(ws) -> tuple[int, dict[int, str], int, int]:
    best: tuple[int, dict[int, str], int, int, int] | None = None
    for row_idx in range(1, min(ws.max_row, 12) + 1):
        headers = {cell.column: _clean_header(cell.value) for cell in ws[row_idx] if _clean_header(cell.value)}
        lower = {col: header.lower() for col, header in headers.items()}
        category_col = next((col for col, h in lower.items() if h in {"equipes", "équipes", "categorie", "catégorie", "pays"}), 1)
        number_col = next((col for col, h in lower.items() if "num" in h or "code" in h), 2)
        person_columns: dict[int, str] = {}
        for col, header in headers.items():
            if col in {category_col, number_col} or _is_ignored_header(header):
                continue
            if col > number_col:
                person_columns[col] = normalize_person_name(header)
        score = len(person_columns) + int(category_col in headers) + int(number_col in headers)
        if best is None or score > best[0]:
            best = (score, person_columns, category_col, number_col, row_idx)
    if not best or not best[1]:
        raise ValueError("Impossible de détecter les colonnes personnes dans l'Excel.")
    _, person_columns, category_col, number_col, row_idx = best
    return row_idx, person_columns, category_col, number_col


def _best_sheet(workbook):
    scored = []
    for ws in workbook.worksheets:
        try:
            header_row, person_columns, category_col, number_col = _find_header_row(ws)
            scored.append((len(person_columns), ws, header_row, person_columns, category_col, number_col))
        except ValueError:
            continue
    if not scored:
        raise ValueError("Aucune feuille compatible trouvée.")
    scored.sort(key=lambda item: item[0], reverse=True)
    return scored[0][1:]


def _make_code(raw_category: str, raw_number: object) -> tuple[str, str, str | None, int | None]:
    number_text = _number_text(raw_number)
    category = normalize_sticker_code(raw_category)
    raw_number_code = normalize_sticker_code(number_text)
    if re.search(r"[A-Za-z]", number_text):
        sticker_code = normalize_sticker_code(number_text)
    elif category == "FWC" and raw_number_code in {"0", "00"}:
        sticker_code = "00"
    else:
        sticker_code = normalize_sticker_code(f"{category}{number_text}")
    category_code, sticker_number = split_sticker_code(sticker_code)
    if category_code is None:
        category_code = category or None
    return sticker_code, make_display_code(sticker_code), category_code, sticker_number


def preview_excel(path_or_file: str | Path | BinaryIO) -> ExcelPreview:
    workbook = load_workbook(path_or_file, data_only=False)
    ws, header_row, person_columns, category_column, number_column = _best_sheet(workbook)
    preview = ExcelPreview(ws.title, header_row, category_column, number_column, person_columns)
    current_category = ""
    album_order = 0
    for row_idx in range(header_row + 1, ws.max_row + 1):
        category_value = ws.cell(row_idx, category_column).value
        number_value = ws.cell(row_idx, number_column).value
        if category_value:
            current_category = _clean_header(category_value)
        if not number_value:
            continue
        if str(current_category).lower().startswith("stickers "):
            continue
        try:
            sticker_code, display_code, category_code, sticker_number = _make_code(current_category, number_value)
        except Exception as exc:
            preview.ignored_rows.append({"row": row_idx, "reason": str(exc)})
            continue
        if not sticker_code:
            preview.ignored_rows.append({"row": row_idx, "reason": "Code sticker vide"})
            continue
        album_order += 1
        quantities = {}
        for col, person_name in person_columns.items():
            cell = ws.cell(row_idx, col)
            if _is_formula_cell(cell):
                quantities[person_name] = 0
            else:
                quantities[person_name] = _as_quantity(cell.value)
        preview.rows.append(
            ParsedStickerRow(
                album_order=album_order,
                raw_category=current_category,
                sticker_code=sticker_code,
                display_code=display_code,
                category_code=category_code,
                sticker_number=sticker_number,
                quantities=quantities,
            )
        )
    return preview


def import_excel(session: Session, path_or_file: str | Path | BinaryIO, source_path: str | None = None) -> dict:
    preview = preview_excel(path_or_file)
    existing_codes = set(session.scalars(select(Sticker.sticker_code)))
    rows_inserted = 0
    rows_updated = 0
    people = {
        name: upsert_person(session, name, display_order=index)
        for index, name in enumerate(preview.people_names, start=1)
    }
    for person in session.scalars(select(Person)):
        if person.name not in people and person.active:
            person.active = False
    for row in preview.rows:
        sticker, created = upsert_sticker(
            session,
            album_order=row.album_order,
            raw_category=row.raw_category,
            category_code=row.category_code,
            category_name=row.raw_category,
            sticker_number=row.sticker_number,
            sticker_code=row.sticker_code,
            display_code=row.display_code,
            source="excel",
        )
        rows_inserted += int(created)
        rows_updated += int(not created and row.sticker_code in existing_codes)
        for person_name, qty in row.quantities.items():
            person = people[person_name]
            holding = get_or_create_holding(session, person.id, sticker.id)
            old_qty = holding.quantity
            holding.quantity = qty
            if old_qty != qty:
                session.add(
                    ActionLog(
                        action_type="import_excel",
                        actor_name=None,
                        person_id=person.id,
                        sticker_id=sticker.id,
                        old_quantity=old_qty,
                        new_quantity=qty,
                        delta=qty - old_qty,
                    )
                )
    import_run = ImportRun(
        import_type="excel",
        source_path=source_path,
        status="success",
        rows_read=preview.sticker_count,
        rows_inserted=rows_inserted,
        rows_updated=rows_updated,
        errors=preview.ignored_rows,
    )
    session.add(import_run)
    session.flush()
    return {
        "import_id": import_run.id,
        "sheet_name": preview.sheet_name,
        "stickers": preview.sticker_count,
        "people": preview.people_count,
        "rows_inserted": rows_inserted,
        "rows_updated": rows_updated,
        "ignored_rows": preview.ignored_rows,
    }


def save_uploaded_file(uploaded_file, target_dir: Path) -> Path:
    target_dir.mkdir(parents=True, exist_ok=True)
    suffix = Path(uploaded_file.name).suffix or ".xlsx"
    with NamedTemporaryFile(delete=False, suffix=suffix, dir=target_dir) as tmp:
        tmp.write(uploaded_file.getbuffer())
        return Path(tmp.name)

